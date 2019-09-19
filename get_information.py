from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def modify_files():
    und_wh="C:\\Users\\admin\\Desktop\\Funko\\script\\und_wh.xlsx"
    und_bo="C:\\Users\\admin\\Desktop\\Funko\\script\\und_bo.xlsx"


    #modifies the warehouse file
    warehouse_wb = load_workbook(filename = und_wh) 
    wh_sheet= warehouse_wb.active
    wh_sheet.delete_cols(7)
    wh_sheet.delete_cols(1)
    wh_sheet.delete_rows(0)
    warehouse_wb.save("und_wh2.xlsx")


    #modifies the backorder file    
    backorder_wb = load_workbook(filename = und_bo)
    bo_sheet = backorder_wb.active

    bo_sheet.delete_cols(10)
    bo_sheet.delete_cols(9) 
    bo_sheet.delete_cols(6)
    bo_sheet.delete_cols(5)
    bo_sheet.delete_cols(3)
    bo_sheet.delete_cols(1)
    bo_sheet.delete_rows(0)
    backorder_wb.save("und_bo2.xlsx")

def delete_zero():
    und_bo="C:\\Users\\admin\\Desktop\\Funko\\script\\und_bo2.xlsx"
    bo_sheet = backorder_wb.active

    for row in bo_sheet.iter_rows():
        for cell in row:
            if cell.value()<=9

    



